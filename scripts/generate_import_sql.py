#!/usr/bin/env python3
"""
Reads Ramps_sample_data_1.xlsx and generates SQL to import into Supabase.
"""

import openpyxl
import datetime
import re

wb = openpyxl.load_workbook("Ramps_sample_data_1.xlsx")


def escape_sql(val):
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"


def format_date(val):
    if val is None:
        return "NULL"
    if isinstance(val, datetime.datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    # Try parsing string dates like "Complete 14/06/2024"
    if isinstance(val, str):
        match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", val)
        if match:
            try:
                d = datetime.datetime.strptime(match.group(1), "%d/%m/%Y")
                return f"'{d.strftime('%Y-%m-%d')}'"
            except ValueError:
                pass
    return "NULL"


def format_num(val):
    if val is None:
        return "0"
    try:
        return str(float(val))
    except (ValueError, TypeError):
        return "0"


# ── Parse Projects sheet ──
ws = wb["Projects"]
projects_data = []
current_school = None

for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=24, values_only=True):
    (
        school_no, name, region, funding_pkg, bundle_id, _,
        pdc, contractor, rbs, _,
        design_complete, construction_complete, admin_complete, dlp_complete, final_closure, _,
        budget, contingency, variations_amt, _,
        defects, _, commentary, _
    ) = row[:24]

    if school_no is not None:
        current_school = {
            "school_no": school_no,
            "name": str(name).strip() if name else None,
            "region": str(region).strip() if region else None,
            "design_complete": design_complete,
            "construction_complete": construction_complete,
            "admin_complete": admin_complete,
            "dlp_complete": dlp_complete,
            "final_closure": final_closure,
            "budget": budget,
            "contingency": contingency,
            "variations_amt": variations_amt,
            "defects": defects,
            "commentary": commentary,
            "ramps": [],
        }
        projects_data.append(current_school)
    elif name and "Ramp" in str(name).strip() and current_school:
        current_school["ramps"].append(str(name).strip())

# ── Parse Status sheet ──
ws2 = wb["Status"]
status_data = {}

for row in ws2.iter_rows(min_row=5, max_row=ws2.max_row, max_col=16, values_only=True):
    (
        _, name, school_no, region, pnn, bundle, pdc, contractor,
        last_milestone, construction_complete_status,
        defects_closed, cfi, school_signoff,
        practical_completion, final_dlp, comments
    ) = row[:16]

    if name:
        status_data[str(name).strip()] = {
            "last_milestone": last_milestone,
            "construction_complete": construction_complete_status,
            "defects_closed": defects_closed,
            "cfi": cfi,
            "school_signoff": school_signoff,
            "practical_completion": practical_completion,
            "final_dlp": final_dlp,
            "comments": comments,
        }


def derive_lifecycle_stage(school, status):
    """Derive lifecycle_stage from Status sheet data."""
    if not status:
        return "construction"

    milestone = status.get("last_milestone")
    final_dlp = status.get("final_dlp")
    practical = status.get("practical_completion")

    # Check milestone text
    if milestone:
        m = str(milestone).strip().lower()
        if "practical completion" in m:
            if final_dlp:
                return "final_closure"
            return "dlp"
        if "construction complete" in m:
            return "admin_closeout"

    # Fallback: check if construction complete date exists
    cc = school.get("construction_complete")
    if cc is not None:
        return "admin_closeout"

    return "construction"


def derive_ramp_status(school, status):
    """Derive ramp status."""
    if not status:
        return "active"
    milestone = status.get("last_milestone")
    if milestone and "practical completion" in str(milestone).lower():
        final_dlp = status.get("final_dlp")
        if final_dlp:
            return "completed"
    return "active"


# ── Generate SQL ──
sql_lines = []

sql_lines.append("-- ============================================")
sql_lines.append("-- VSBA RAMPS Data Import from Excel")
sql_lines.append("-- Generated from Ramps_sample_data_1.xlsx")
sql_lines.append("-- ============================================")
sql_lines.append("")

# Step 1: Rename regions to abbreviations
sql_lines.append("-- Step 1: Rename regions to abbreviations")
sql_lines.append("UPDATE regions SET name = 'NEV' WHERE name = 'North-Eastern Victoria';")
sql_lines.append("UPDATE regions SET name = 'NWV' WHERE name = 'North-Western Victoria';")
sql_lines.append("UPDATE regions SET name = 'SEV' WHERE name = 'South-Eastern Victoria';")
sql_lines.append("UPDATE regions SET name = 'SWV' WHERE name = 'South-Western Victoria';")
sql_lines.append("")

# Step 2: Clear existing seed data (cascades to ramps, milestones, etc.)
sql_lines.append("-- Step 2: Clear existing seed data")
sql_lines.append("DELETE FROM communications;")
sql_lines.append("DELETE FROM schools;")
sql_lines.append("")

# Step 3: Insert schools and ramps
sql_lines.append("-- Step 3: Insert schools and ramps")
sql_lines.append("DO $$")
sql_lines.append("DECLARE")
sql_lines.append("  v_program_id UUID;")
sql_lines.append("  v_region_id UUID;")
sql_lines.append("  v_school_id UUID;")
sql_lines.append("  v_ramp_id UUID;")
sql_lines.append("BEGIN")
sql_lines.append("  SELECT id INTO v_program_id FROM programs WHERE name = 'RAMPS Program' LIMIT 1;")
sql_lines.append("")

for school in projects_data:
    if not school["name"] or not school["region"]:
        continue

    status = status_data.get(school["name"], {})
    lifecycle = derive_lifecycle_stage(school, status)
    ramp_status = derive_ramp_status(school, status)

    # Determine school status
    if ramp_status == "completed":
        school_status = "completed"
    else:
        school_status = "active"

    # Build commentary from Projects commentary + Status comments
    commentary_parts = []
    if school["commentary"]:
        commentary_parts.append(str(school["commentary"]))
    if status and status.get("comments") and str(status.get("comments", "")) != str(school.get("commentary", "")):
        status_comments = str(status["comments"])
        if status_comments not in commentary_parts:
            commentary_parts.append(status_comments)
    combined_commentary = "\n\n".join(commentary_parts) if commentary_parts else None

    sql_lines.append(f"  -- {school['name']}")
    sql_lines.append(f"  SELECT id INTO v_region_id FROM regions WHERE name = {escape_sql(school['region'])} LIMIT 1;")
    sql_lines.append(f"  INSERT INTO schools (id, program_id, region_id, name, status)")
    sql_lines.append(f"    VALUES (gen_random_uuid(), v_program_id, v_region_id, {escape_sql(school['name'])}, '{school_status}')")
    sql_lines.append(f"    RETURNING id INTO v_school_id;")

    # Determine ramps to create
    if school["ramps"]:
        ramp_names = school["ramps"]
    else:
        ramp_names = ["Ramp 1"]

    for ramp_name in ramp_names:
        budget = format_num(school["budget"])
        actual = format_num(school["budget"])  # actual = budget from the sheet (TOTAL Approved FUNDING)
        forecast = format_num(school["contingency"])

        sql_lines.append(f"  INSERT INTO ramps (id, school_id, name, lifecycle_stage, status, commentary, budget_amount, actual_amount, forecast_amount)")
        sql_lines.append(f"    VALUES (gen_random_uuid(), v_school_id, {escape_sql(ramp_name)}, '{lifecycle}', '{ramp_status}', {escape_sql(combined_commentary)}, {budget}, {actual}, {forecast})")
        sql_lines.append(f"    RETURNING id INTO v_ramp_id;")

        # Update auto-seeded milestones with dates from the Projects sheet
        if school["design_complete"]:
            d = format_date(school["design_complete"])
            if d != "NULL":
                sql_lines.append(f"  UPDATE milestones SET actual_date = {d} WHERE ramp_id = v_ramp_id AND name = 'Design Signoff';")

        if school["construction_complete"]:
            d = format_date(school["construction_complete"])
            if d != "NULL":
                sql_lines.append(f"  UPDATE milestones SET actual_date = {d} WHERE ramp_id = v_ramp_id AND name = 'Practical Completion';")

        if school["admin_complete"]:
            d = format_date(school["admin_complete"])
            if d != "NULL":
                sql_lines.append(f"  UPDATE milestones SET actual_date = {d} WHERE ramp_id = v_ramp_id AND name = 'Administrative Completion';")

        if school["dlp_complete"]:
            d = format_date(school["dlp_complete"])
            if d != "NULL":
                sql_lines.append(f"  UPDATE milestones SET actual_date = {d} WHERE ramp_id = v_ramp_id AND name = 'DLP Complete';")

        if school["final_closure"]:
            d = format_date(school["final_closure"])
            if d != "NULL":
                sql_lines.append(f"  UPDATE milestones SET actual_date = {d} WHERE ramp_id = v_ramp_id AND name = 'Final Closure';")

    sql_lines.append("")

sql_lines.append("END $$;")
sql_lines.append("")
sql_lines.append(f"-- Import complete: {len(projects_data)} schools")

output = "\n".join(sql_lines)

with open("scripts/import_ramps_data.sql", "w") as f:
    f.write(output)

print(f"Generated SQL: {len(projects_data)} schools, {sum(max(1, len(s['ramps'])) for s in projects_data)} ramps")
print(f"Written to scripts/import_ramps_data.sql")
